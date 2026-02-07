"""Serviço de exportação de relatórios em PDF e Excel."""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from repositories.entrada_repo import EntradaRepository
from repositories.saida_repo import SaidaRepository
from repositories.quebrado_repo import QuebradoRepository
from repositories.despesa_repo import DespesaRepository
from repositories.resumo_repo import ResumoRepository


class ExportService:
    """Gera relatórios em PDF e Excel."""

    MESES_PT = [
        'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
        'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
    ]

    @staticmethod
    def _nome_mes(mes_ref):
        """Converte '2026-02' para 'Fevereiro 2026'."""
        ano, mes = mes_ref.split('-')
        return f"{ExportService.MESES_PT[int(mes) - 1]} {ano}"

    @staticmethod
    def exportar_excel(mes_referencia):
        """
        Gera um arquivo Excel com o relatório do mês.

        Returns:
            BytesIO com o conteúdo do arquivo .xlsx
        """
        resumo = ResumoRepository.get_by_month(mes_referencia)
        entradas = EntradaRepository.get_by_month(mes_referencia)
        saidas = SaidaRepository.get_by_month(mes_referencia)
        quebrados = QuebradoRepository.get_by_month(mes_referencia)
        despesas = DespesaRepository.get_by_month(mes_referencia)

        wb = Workbook()
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        title_font = Font(bold=True, size=14, color='1E293B')
        border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0'),
        )

        # ── Aba: Resumo ──
        ws = wb.active
        ws.title = 'Resumo'
        ws.sheet_properties.tabColor = '4F46E5'

        ws.merge_cells('A1:D1')
        ws['A1'] = f'🥚 EggVault — Relatório {ExportService._nome_mes(mes_referencia)}'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A3'] = 'Indicador'
        ws['B3'] = 'Valor'
        for col in ['A', 'B']:
            ws[f'{col}3'].font = header_font
            ws[f'{col}3'].fill = header_fill
            ws[f'{col}3'].alignment = Alignment(horizontal='center')

        dados_resumo = [
            ('Total de Entradas', resumo['total_entradas']),
            ('Total de Saídas (Vendas)', resumo['total_saidas']),
            ('Total de Quebrados', resumo.get('total_quebrados', 0)),
            ('Faturamento Total', f"R$ {resumo['faturamento_total']:.2f}"),
            ('Total Despesas', f"R$ {resumo.get('total_despesas', 0):.2f}"),
            ('Lucro Líquido', f"R$ {resumo.get('lucro_estimado', 0):.2f}"),
            ('Saldo do Mês (ovos)', resumo['total_entradas'] - resumo['total_saidas'] - resumo.get('total_quebrados', 0)),
        ]

        for i, (label, valor) in enumerate(dados_resumo, start=4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = valor
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'A{i}'].border = border
            ws[f'B{i}'].border = border
            ws[f'B{i}'].alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        # ── Aba: Entradas ──
        ws_ent = wb.create_sheet('Entradas')
        ws_ent.sheet_properties.tabColor = '10B981'
        headers_ent = ['Data', 'Quantidade', 'Observação']
        for j, h in enumerate(headers_ent, 1):
            cell = ws_ent.cell(row=1, column=j, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        for i, e in enumerate(entradas, start=2):
            ws_ent.cell(row=i, column=1, value=e['data']).border = border
            ws_ent.cell(row=i, column=2, value=e['quantidade']).border = border
            ws_ent.cell(row=i, column=3, value=e.get('observacao', '')).border = border

        ws_ent.column_dimensions['A'].width = 22
        ws_ent.column_dimensions['B'].width = 14
        ws_ent.column_dimensions['C'].width = 35

        # ── Aba: Vendas ──
        ws_ven = wb.create_sheet('Vendas')
        ws_ven.sheet_properties.tabColor = 'EF4444'
        headers_ven = ['Data', 'Quantidade', 'Preço Unit.', 'Valor Total']
        for j, h in enumerate(headers_ven, 1):
            cell = ws_ven.cell(row=1, column=j, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        for i, s in enumerate(saidas, start=2):
            ws_ven.cell(row=i, column=1, value=s['data']).border = border
            ws_ven.cell(row=i, column=2, value=s['quantidade']).border = border
            ws_ven.cell(row=i, column=3, value=f"R$ {s['preco_unitario']:.2f}").border = border
            ws_ven.cell(row=i, column=4, value=f"R$ {s['valor_total']:.2f}").border = border

        ws_ven.column_dimensions['A'].width = 22
        ws_ven.column_dimensions['B'].width = 14
        ws_ven.column_dimensions['C'].width = 14
        ws_ven.column_dimensions['D'].width = 16

        # ── Aba: Quebrados ──
        ws_q = wb.create_sheet('Quebrados')
        ws_q.sheet_properties.tabColor = 'BE185D'
        headers_q = ['Data', 'Quantidade', 'Motivo']
        for j, h in enumerate(headers_q, 1):
            cell = ws_q.cell(row=1, column=j, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color='BE185D', end_color='BE185D', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        for i, q in enumerate(quebrados, start=2):
            ws_q.cell(row=i, column=1, value=q['data']).border = border
            ws_q.cell(row=i, column=2, value=q['quantidade']).border = border
            ws_q.cell(row=i, column=3, value=q.get('motivo', '')).border = border

        ws_q.column_dimensions['A'].width = 22
        ws_q.column_dimensions['B'].width = 14
        ws_q.column_dimensions['C'].width = 35

        # ── Aba: Despesas ──
        ws_d = wb.create_sheet('Despesas')
        ws_d.sheet_properties.tabColor = 'C2410C'
        headers_d = ['Data', 'Valor', 'Descrição']
        for j, h in enumerate(headers_d, 1):
            cell = ws_d.cell(row=1, column=j, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color='C2410C', end_color='C2410C', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        for i, d in enumerate(despesas, start=2):
            ws_d.cell(row=i, column=1, value=d['data']).border = border
            ws_d.cell(row=i, column=2, value=f"R$ {d['valor']:.2f}").border = border
            ws_d.cell(row=i, column=3, value=d.get('descricao', '')).border = border

        ws_d.column_dimensions['A'].width = 22
        ws_d.column_dimensions['B'].width = 16
        ws_d.column_dimensions['C'].width = 35

        # Salvar em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def exportar_pdf(mes_referencia):
        """
        Gera um arquivo PDF com o relatório do mês.

        Returns:
            BytesIO com o conteúdo do arquivo .pdf
        """
        resumo = ResumoRepository.get_by_month(mes_referencia)
        entradas = EntradaRepository.get_by_month(mes_referencia)
        saidas = SaidaRepository.get_by_month(mes_referencia)
        quebrados = QuebradoRepository.get_by_month(mes_referencia)
        despesas = DespesaRepository.get_by_month(mes_referencia)

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Title'],
            fontSize=18, textColor=colors.HexColor('#1E293B'),
            spaceAfter=6
        )
        subtitle_style = ParagraphStyle(
            'CustomSubtitle', parent=styles['Normal'],
            fontSize=10, textColor=colors.HexColor('#64748B'),
            spaceAfter=20
        )
        section_style = ParagraphStyle(
            'SectionTitle', parent=styles['Heading2'],
            fontSize=13, textColor=colors.HexColor('#4F46E5'),
            spaceAfter=8, spaceBefore=16
        )

        # Título
        nome_mes = ExportService._nome_mes(mes_referencia)
        elements.append(Paragraph(f'🥚 EggVault — Relatório {nome_mes}', title_style))
        elements.append(Paragraph(f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}', subtitle_style))

        # ── Resumo ──
        elements.append(Paragraph('Resumo do Mês', section_style))

        saldo = resumo['total_entradas'] - resumo['total_saidas'] - resumo.get('total_quebrados', 0)
        resumo_data = [
            ['Indicador', 'Valor'],
            ['Total de Entradas', str(resumo['total_entradas'])],
            ['Total de Saídas', str(resumo['total_saidas'])],
            ['Total Quebrados', str(resumo.get('total_quebrados', 0))],
            ['Faturamento', f"R$ {resumo['faturamento_total']:.2f}"],
            ['Total Despesas', f"R$ {resumo.get('total_despesas', 0):.2f}"],
            ['Lucro Líquido', f"R$ {resumo.get('lucro_estimado', 0):.2f}"],
            ['Saldo (ovos)', str(saldo)],
        ]
        t = Table(resumo_data, colWidths=[120*mm, 50*mm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(t)

        # ── Entradas ──
        if entradas:
            elements.append(Paragraph(f'Entradas ({len(entradas)} registros)', section_style))
            ent_data = [['Data', 'Quantidade', 'Observação']]
            for e in entradas:
                data_fmt = datetime.fromisoformat(e['data']).strftime('%d/%m/%Y %H:%M') if e.get('data') else '—'
                ent_data.append([data_fmt, str(e['quantidade']), e.get('observacao', '') or '—'])

            t = Table(ent_data, colWidths=[45*mm, 30*mm, 95*mm])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0FDF4')]),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(t)

        # ── Vendas ──
        if saidas:
            elements.append(Paragraph(f'Vendas ({len(saidas)} registros)', section_style))
            ven_data = [['Data', 'Qtd', 'Preço Unit.', 'Total']]
            for s in saidas:
                data_fmt = datetime.fromisoformat(s['data']).strftime('%d/%m/%Y %H:%M') if s.get('data') else '—'
                ven_data.append([
                    data_fmt, str(s['quantidade']),
                    f"R$ {s['preco_unitario']:.2f}", f"R$ {s['valor_total']:.2f}"
                ])

            t = Table(ven_data, colWidths=[45*mm, 25*mm, 35*mm, 35*mm])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FEF2F2')]),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ]))
            elements.append(t)

        # ── Quebrados ──
        if quebrados:
            elements.append(Paragraph(f'Quebrados ({len(quebrados)} registros)', section_style))
            q_data = [['Data', 'Quantidade', 'Motivo']]
            for q in quebrados:
                data_fmt = datetime.fromisoformat(q['data']).strftime('%d/%m/%Y %H:%M') if q.get('data') else '—'
                q_data.append([data_fmt, str(q['quantidade']), q.get('motivo', '') or '—'])

            t = Table(q_data, colWidths=[45*mm, 30*mm, 95*mm])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#BE185D')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FCE7F3')]),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(t)

        # ── Despesas ──
        if despesas:
            elements.append(Paragraph(f'Despesas ({len(despesas)} registros)', section_style))
            d_data = [['Data', 'Valor', 'Descrição']]
            for d in despesas:
                data_fmt = datetime.fromisoformat(d['data']).strftime('%d/%m/%Y %H:%M') if d.get('data') else '—'
                d_data.append([data_fmt, f"R$ {d['valor']:.2f}", d.get('descricao', '') or '—'])

            t = Table(d_data, colWidths=[45*mm, 30*mm, 95*mm])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C2410C')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF7ED')]),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(t)

        doc.build(elements)
        output.seek(0)
        return output

    @staticmethod
    def exportar_excel_anual(ano):
        """
        Gera Excel com resumo de todos os meses do ano.

        Returns:
            BytesIO com o conteúdo do arquivo .xlsx
        """
        resumos = ResumoRepository.get_by_year(ano)

        wb = Workbook()
        ws = wb.active
        ws.title = f'Resumo {ano}'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0'),
        )

        title_font = Font(bold=True, size=14, color='1E293B')
        ws.merge_cells('A1:F1')
        ws['A1'] = f'🥚 EggVault — Resumo Anual {ano}'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        headers = ['Mês', 'Entradas', 'Saídas', 'Quebrados', 'Faturamento', 'Despesas', 'Lucro Líquido', 'Saldo']
        for j, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=j, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for i, r in enumerate(resumos, start=4):
            nome = ExportService._nome_mes(r['mes_referencia'])
            saldo = r['total_entradas'] - r['total_saidas'] - r.get('total_quebrados', 0)
            ws.cell(row=i, column=1, value=nome).border = border
            ws.cell(row=i, column=2, value=r['total_entradas']).border = border
            ws.cell(row=i, column=3, value=r['total_saidas']).border = border
            ws.cell(row=i, column=4, value=r.get('total_quebrados', 0)).border = border
            ws.cell(row=i, column=5, value=f"R$ {r['faturamento_total']:.2f}").border = border
            ws.cell(row=i, column=6, value=f"R$ {r.get('total_despesas', 0):.2f}").border = border
            ws.cell(row=i, column=7, value=f"R$ {r.get('lucro_estimado', 0):.2f}").border = border
            ws.cell(row=i, column=8, value=saldo).border = border

        for col, w in [('A', 20), ('B', 12), ('C', 12), ('D', 12), ('E', 16), ('F', 16), ('G', 16), ('H', 10)]:
            ws.column_dimensions[col].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
